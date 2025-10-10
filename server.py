import os
import io
import re
from datetime import datetime
from flask import Flask, request, jsonify, send_from_directory, abort, Response
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PUBLIC_DIR = os.path.join(BASE_DIR, "public")

app = Flask(__name__, static_folder=PUBLIC_DIR, static_url_path='')

# ---------- STATIC ----------
@app.get("/")
def index():
    index_path = os.path.join(PUBLIC_DIR, "index.html")
    if not os.path.exists(index_path):
        return "index.html non trovato in /public", 404
    return send_from_directory(PUBLIC_DIR, "index.html")

@app.get("/<path:asset>")
def static_files(asset):
    path = os.path.join(PUBLIC_DIR, asset)
    if not os.path.exists(path):
        abort(404)
    directory, fname = os.path.split(path)
    return send_from_directory(directory, fname)

# ---------- FILE API (GET/PUT) ----------
@app.get("/api/files/<path:name>")
def api_read_file(name):
    safe = secure_filename(name)
    fpath = os.path.join(PUBLIC_DIR, safe)
    if not os.path.exists(fpath):
        return "", 404
    with open(fpath, "rb") as f:
        data = f.read()
    resp = Response(data, mimetype="text/plain; charset=utf-8")
    # important to avoid stale caches while testing in the browser
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    return resp

@app.put("/api/files/<path:name>")
def api_write_file(name):
    safe = secure_filename(name)
    fpath = os.path.join(PUBLIC_DIR, safe)
    os.makedirs(os.path.dirname(fpath) or PUBLIC_DIR, exist_ok=True)
    data = request.get_data() or b""
    with open(fpath, "wb") as f:
        f.write(data)
    return jsonify(ok=True, path=safe)

# ---------- UTILITIES EXCEL ----------
DATE_RE = re.compile(r"^(\d{2})-(\d{2})-(\d{4})$")
MONTHS_IT = ["Gen","Feb","Mar","Apr","Mag","Giu","Lug","Ago","Set","Ott","Nov","Dic"]

def col_letter(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n-1, 26)
        s = chr(65 + r) + s
    return s

def _best_col_width(values, minimum=10, maximum=35, padding=2):
    max_len = 0
    for v in values:
        if v is None:
            continue
        s = str(v)
        l = sum(1.2 if c.isupper() else 1 for c in s)
        if l > max_len:
            max_len = l
    width = max(minimum, min(maximum, int(max_len + padding)))
    return width

def build_workbook_from_sheets(sheets: dict) -> Workbook:
    """
    sheets: { "Gen":[["Data","A","B"],["01-01-2025","TS1",""]], ... }
    Ordina i fogli dal mese corrente (oggi) e poi a scalare (wrap).
    Applica formattazione: header bold + fill, bordi, centering, freeze panes, larghezze.
    """
    today = datetime.now()
    start_m = today.month
    order = list(range(start_m, 13)) + list(range(1, start_m))
    ordered_keys = []
    for m in order:
        key = MONTHS_IT[m-1]
        if key in sheets:
            ordered_keys.append(key)
    for k in sheets.keys():
        if k not in ordered_keys:
            ordered_keys.append(k)

    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)

    thin = Side(border_style="thin", color="D0D7E2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="EEF3FA")
    head_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for key in ordered_keys:
        ws = wb.create_sheet(title=key[:31])
        rows = sheets.get(key, [])
        # Write rows
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.alignment = center
                cell.border = border
                if r_idx == 1:
                    cell.fill = head_fill
                    cell.font = head_font

        # Freeze panes (keep header visible)
        ws.freeze_panes = "A2"

        # Auto width
        last_col = max((len(r) for r in rows), default=1)
        columns_values = {c: [] for c in range(1, last_col + 1)}
        for r in range(1, len(rows) + 1):
            for c in range(1, last_col + 1):
                columns_values[c].append(ws.cell(row=r, column=c).value)

        for c in range(1, last_col + 1):
            col = col_letter(c)
            if c == 1:
                ws.column_dimensions[col].width = 12
            else:
                width = _best_col_width(columns_values[c], minimum=10, maximum=30, padding=2)
                ws.column_dimensions[col].width = width

        ws.auto_filter.ref = ws.dimensions

    return wb

# ---------- EXPORT XLSX (download) ----------
@app.post("/api/export/xlsx")
def api_export_xlsx():
    payload = request.get_json(force=True, silent=True)
    if not payload:
        return jsonify(ok=False, error="JSON mancante"), 400
    sheets = payload.get("sheets")
    if not isinstance(sheets, dict) or not sheets:
        return jsonify(ok=False, error="'sheets' mancante o vuoto"), 400

    filename = payload.get("filename") or f"monthly_data_{datetime.now().year}.xlsx"
    safe_name = secure_filename(filename) or "export.xlsx"

    wb = build_workbook_from_sheets(sheets)
    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)

    resp = Response(
        mem.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp.headers["Content-Disposition"] = f'attachment; filename="{safe_name}"'
    return resp

# ---------- EXPORT XLSX (SALVA SU DISCO) ----------
@app.post("/api/export/xlsx/save")
def api_export_xlsx_save():
    payload = request.get_json(force=True, silent=True)
    if not payload:
        return jsonify(ok=False, error="JSON mancante"), 400
    sheets = payload.get("sheets")
    if not isinstance(sheets, dict) or not sheets:
        return jsonify(ok=False, error="'sheets' mancante o vuoto"), 400

    filename = payload.get("filename") or f"monthly_data_{datetime.now().year}.xlsx"
    safe_name = secure_filename(filename) or "export.xlsx"
    out_path = os.path.join(PUBLIC_DIR, safe_name)

    wb = build_workbook_from_sheets(sheets)
    try:
        wb.save(out_path)
    except Exception as e:
        return jsonify(ok=False, error=f"write_error: {e}"), 500
    return jsonify(ok=True, path=safe_name)

# ---------- DUAL YEAR HELPERS ----------
_sel_re = re.compile(r"^selections_(\d{4})\.txt$", re.IGNORECASE)

def _list_selection_years():
    years = []
    try:
        for name in os.listdir(PUBLIC_DIR):
            m = _sel_re.match(name)
            if m:
                y = int(m.group(1))
                years.append(y)
    except Exception:
        pass
    years = sorted(list(set(years)))
    return years

@app.get("/api/years")
def api_years():
    years = _list_selection_years()
    now = datetime.now().year
    suggested = []
    if (now in years) and ((now+1) in years):
        suggested = [now, now+1]
    elif len(years) >= 2:
        suggested = years[-2:]
    elif len(years) == 1:
        suggested = [years[0]]
    return jsonify(ok=True, years=years, suggested_pair=suggested)

@app.get("/api/selections/<int:year>")
def api_get_selections(year: int):
    name = f"selections_{year}.txt"
    path = os.path.join(PUBLIC_DIR, name)
    if not os.path.exists(path):
        return "", 404
    with open(path, "rb") as f:
        data = f.read()
    resp = Response(data, mimetype="text/plain; charset=utf-8")
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    return resp

@app.get("/api/selections/combined")
def api_get_selections_combined():
    qs = request.args.get("years") or ""
    years = []
    for part in re.split(r"[,\\s]+", qs.strip()):
        try:
            y = int(part)
            if y not in years:
                years.append(y)
        except Exception:
            pass
    years = years[:2]
    if not years:
        return "", 400

    chunks = []
    for y in years:
        name = f"selections_{y}.txt"
        path = os.path.join(PUBLIC_DIR, name)
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    chunks.append(f.read())
            except Exception:
                pass
    data = ("\n".join(chunks)).strip()
    resp = Response(data, mimetype="text/plain; charset=utf-8")
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    return resp

if __name__ == "__main__":
    os.makedirs(PUBLIC_DIR, exist_ok=True)
    app.run(host="127.0.0.1", port=8000, debug=True)
